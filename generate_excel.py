"""
QA Test Case Generator — Excel Output
======================================
Workflow input -> Groq AI -> Styled Excel (.xlsx) with pill-badge dropdowns.

Required Environment Variables (set in .env):
    GROQ_API_KEY   - Your free Groq API key
    GROQ_MODEL     - Groq model (default: llama-3.3-70b-versatile)

Usage:
    python generate_excel.py
"""

import os
import re
import sys
from datetime import date

# Load .env file
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except ImportError:
    pass

from groq import Groq  # type: ignore
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

HEADERS = [
    "Test Case ID",   # A
    "Test Scenario",  # B
    "Test Steps",     # C
    "Expected Result",# D
    "Actual Result",  # E
    "Test Case Type", # F  — AI-generated
    "Priority",       # G  — AI-generated
    "Status",         # H
    "Date",           # I
    "Completed Date", # J
    "Tested By",      # K
    "Comments",       # L
]

# Column widths (characters)
COL_WIDTHS = {
    "A": 14,   # Test Case ID
    "B": 30,   # Test Scenario
    "C": 45,   # Test Steps
    "D": 35,   # Expected Result
    "E": 25,   # Actual Result
    "F": 18,   # Test Case Type
    "G": 12,   # Priority
    "H": 12,   # Status
    "I": 14,   # Date
    "J": 16,   # Completed Date
    "K": 18,   # Tested By
    "L": 30,   # Comments
}

# ---------------------------------------------------------------------------
# Pill color palettes
# ---------------------------------------------------------------------------

# Test Case Type  →  (bg_hex, font_hex)
TC_TYPE_PILLS = {
    "Functional": ("FF8C00", "FFFFFF"),   # Orange bg, white text
    "Positive":   ("28A645", "FFFFFF"),   # Green bg, white text
    "Negative":   ("DC3545", "FFFFFF"),   # Red bg, white text
    "UI/UX":      ("0078D4", "FFFFFF"),   # Blue bg, white text
    "Security":   ("6F42C1", "FFFFFF"),   # Purple bg, white text
}

# Priority  →  (bg_hex, font_hex)
PRIORITY_PILLS = {
    "High":   ("DC3545", "FFFFFF"),   # Red bg, white text
    "Medium": ("FF8C00", "FFFFFF"),   # Orange bg, white text
    "Low":    ("28A645", "FFFFFF"),   # Green bg, white text
}

# Status  →  (bg_hex, font_hex)
STATUS_PILLS = {
    "Pass": ("28A645", "FFFFFF"),   # Green bg, white text
    "Fail": ("DC3545", "FFFFFF"),   # Red bg, white text
    "NS":   ("6C757D", "FFFFFF"),   # Grey bg, white text (Not Started)
}

# Header styling
HEADER_BG   = "1B4332"   # Dark green
HEADER_FONT = "FFFFFF"   # White


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def pill_font(hex_color: str, bold: bool = True, size: int = 10) -> Font:
    return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")


def thin_border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def center_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def apply_pill_conditional_formatting(ws, col_letter: str, pill_map: dict, max_row: int = 1000):
    """
    Apply conditional formatting to simulate pill badges.
    For each value in pill_map, when cell equals that value:
      - background is set to pill color
      - text colour is set accordingly
    """
    for value, (bg_hex, fg_hex) in pill_map.items():
        fill = make_fill(bg_hex)
        font = pill_font(fg_hex)
        ds   = DifferentialStyle(fill=fill, font=font)
        rule = Rule(
            type="containsText",
            operator="containsText",
            text=value,
            dxf=ds,
        )
        rule.formula = [f'NOT(ISERROR(SEARCH("{value}",{col_letter}2)))']
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)


def add_dropdown(ws, col_letter: str, options: list, max_row: int = 1000):
    """Add a dropdown validation to an entire column (skipping header row 1)."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        showDropDown=False,   # False = SHOW the dropdown arrow
        showErrorMessage=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}2:{col_letter}{max_row}"


# ---------------------------------------------------------------------------
# Groq: generate test cases
# ---------------------------------------------------------------------------

def generate_test_cases(workflow: str) -> list[list[str]]:
    """Call Groq API and return parsed test case rows (5 fields each)."""
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "Missing env var: GROQ_API_KEY\n"
            "Get a free key from https://console.groq.com/keys"
        )

    client = Groq(api_key=api_key)

    system_prompt = (
        "You are a Senior QA Automation Engineer. "
        "You generate comprehensive, exhaustive, multi-level QA test cases strictly in pipe-separated format. "
        "No markdown, no headers, no explanations — only test case lines."
    )

    user_prompt = f"""You are a Senior QA Automation Engineer.

Generate comprehensive and detailed QA test cases for the following workflow.

WORKFLOW:
{workflow}

IMPORTANT REQUIREMENTS:
- Generate test cases at ALL levels: Dashboard, Module, Subfolder, and UI Component.
- For EACH module and subfolder, cover: Navigation, Page loading, UI visibility, Button functionality,
  Form validation, Table data validation, Search, Filter, Sort, Create, Edit, Delete,
  Export, Error handling, Permission validation, Negative scenarios, and Edge cases.
- Go deep inside EACH subfolder with detailed component-level test cases.
- There is NO LIMIT on the number of test cases. Generate as many as required.

OUTPUT FORMAT RULES:
- Each test case on its own line.
- Output EXACTLY 6 pipe "|" separated fields:
  TC_ID | Test Scenario | Test Steps | Expected Result | Test Case Type | Priority
- Test Case IDs must be sequential and unique: TC001, TC002, TC003, ...
- Test Scenario: Natural, action-oriented. Do NOT start with the word "Verify".
- Test Steps: Specific actions separated by semicolons.
- Expected Result: Precise, verifiable outcome.
- Test Case Type: EXACTLY one of — Functional, Negative, UI/UX, Validation
- Priority: EXACTLY one of — High, Medium, Low
- No header row. No markdown. No bulleted lists.

EXAMPLE OUTPUT:
TC001 | Successful login with valid credentials | Navigate to login page; Enter valid username and password; Click Login button | User is redirected to the main dashboard | Functional | High
TC002 | Login attempt with incorrect password | Navigate to login page; Enter valid username; Enter wrong password; Click Login | Error message displayed; User stays on login page | Negative | High
TC003 | Dashboard loading time within threshold | Open dashboard; Measure time until all widgets load | Dashboard loads within 3 seconds | UI/UX | Medium
TC004 | Asset list table column header sorting | Navigate to Asset List; Click on Name column header | Table rows sort alphabetically; Sort arrow indicator visible | Functional | Medium
TC005 | Add asset form required field validation | Navigate to Add Asset form; Leave all fields blank; Click Submit | Validation errors shown for all required fields; Form not submitted | Validation | High

Now generate exhaustive test cases for the workflow above:"""

    model = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
    print(f"[Groq] Sending workflow to {model}...")

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_prompt},
            ],
            temperature=0.2,
        )
    except Exception as e:
        raise Exception(f"[Groq] API call failed: {e}")

    raw = response.choices[0].message.content.strip()
    if not raw:
        raise ValueError("[Groq] Empty response received.")

    print("[Groq] Response received. Parsing...")

    test_cases = []
    for line_num, line in enumerate(raw.splitlines(), start=1):
        line = line.strip()
        if not line or line.startswith("```") or line.startswith("#"):
            continue

        parts = [p.strip() for p in line.split("|")]
        if len(parts) != 6:
            print(f"  [Skip] Line {line_num}: expected 6 fields, got {len(parts)}")
            continue

        if not re.match(r"^TC\d+$", parts[0], re.IGNORECASE):
            print(f"  [Skip] Line {line_num}: invalid TC ID '{parts[0]}'")
            continue

        # Format steps as numbered list
        steps = parts[2]
        items = [s.strip() for s in steps.split(";") if s.strip()]
        if items:
            parts[2] = "\n".join(f"{i}. {s}" for i, s in enumerate(items, 1))

        test_cases.append(parts)  # 6 fields: ID, Scenario, Steps, Expected, TCType, Priority

    print(f"[Groq] Parsed {len(test_cases)} test case(s).")
    return test_cases


# ---------------------------------------------------------------------------
# Build the Excel workbook
# ---------------------------------------------------------------------------

def build_workbook(test_cases: list[list[str]], output_path: str) -> None:
    """Create a styled Excel workbook with pill-badge dropdowns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Column widths ────────────────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Header row ───────────────────────────────────────────────────────────
    header_fill = make_fill(HEADER_BG)
    header_font = Font(color=HEADER_FONT, bold=True, size=11, name="Segoe UI")
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = center_align()
        cell.border  = thin_border()

    # ── Add dropdowns ─────────────────────────────────────────────────────────
    add_dropdown(ws, "F", list(TC_TYPE_PILLS.keys()))    # Test Case Type
    add_dropdown(ws, "G", list(PRIORITY_PILLS.keys()))   # Priority
    add_dropdown(ws, "H", list(STATUS_PILLS.keys()))     # Status

    # ── Conditional formatting (pill colors) ──────────────────────────────────
    apply_pill_conditional_formatting(ws, "F", TC_TYPE_PILLS)
    apply_pill_conditional_formatting(ws, "G", PRIORITY_PILLS)
    apply_pill_conditional_formatting(ws, "H", STATUS_PILLS)

    # ── Write data rows ───────────────────────────────────────────────────────
    today_str = date.today().strftime("%d/%m/%Y")
    border    = thin_border()

    for row_idx, tc in enumerate(test_cases, start=2):
        tc_id, scenario, steps, expected, tc_type, priority = tc

        # Build full 12-value row
        row_data = [
            tc_id,        # A: Test Case ID
            scenario,     # B: Test Scenario
            steps,        # C: Test Steps
            expected,     # D: Expected Result
            "",           # E: Actual Result (blank)
            tc_type,      # F: Test Case Type (AI-generated)
            priority,     # G: Priority (AI-generated)
            "",           # H: Status (blank)
            today_str,    # I: Date
            "",           # J: Completed Date (blank)
            "",           # K: Tested By (blank)
            "",           # L: Comments (blank)
        ]

        # Estimate height from number of step lines
        step_lines = max(steps.count("\n") + 1, 1)
        ws.row_dimensions[row_idx].height = max(20, step_lines * 15 + 5)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            col_letter = get_column_letter(col_idx)

            # Alignment
            if col_letter in ("C",):       # Test Steps — wrap, left
                cell.alignment = left_align(wrap=True)
            elif col_letter in ("F", "G", "H"):   # Pill columns — center
                cell.alignment = center_align()
                cell.font = Font(bold=True, size=10, name="Segoe UI")
            elif col_letter in ("A", "I", "J"):  # ID & Dates — center
                cell.alignment = center_align()
                cell.font = Font(size=10, name="Segoe UI")
            elif col_letter in ("K", "L"):  # Tested By / Comments — left
                cell.alignment = left_align(wrap=True)
                cell.font = Font(size=10, name="Segoe UI")
            else:
                cell.alignment = left_align(wrap=True)
                cell.font = Font(size=10, name="Segoe UI")

            # Zebra row shading (light grey on even rows)
            if row_idx % 2 == 0 and col_letter not in ("F", "G", "H"):
                cell.fill = make_fill("F7F7F7")

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"[Excel] Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  QA Test Case Generator — Excel Output")
    print(f"  Model: {GROQ_MODEL}")
    print("=" * 60)

    # Workflow input
    print("\nEnter your workflow description below.")
    print("Press Enter twice when finished.\n")

    lines = []
    try:
        while True:
            line = input()
            if line == "" and lines and lines[-1] == "":
                break
            lines.append(line)
    except KeyboardInterrupt:
        print("\n\n[Info] Cancelled.")
        sys.exit(0)

    workflow = "\n".join(lines).strip()
    if not workflow:
        print("\n[Error] Workflow cannot be empty.")
        sys.exit(1)

    print(f"\n[Info] Workflow captured ({len(workflow)} chars).\n")

    # Generate test cases
    try:
        test_cases = generate_test_cases(workflow)
    except Exception as e:
        print(f"\n[Error] {e}")
        sys.exit(1)

    # Build Excel file
    output_file = f"test_cases_{date.today().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), output_file)

    try:
        build_workbook(test_cases, output_path)
    except Exception as e:
        print(f"\n[Error] Failed to create Excel: {e}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print(f"  SUCCESS! {len(test_cases)} test cases saved.")
    print(f"  File: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
